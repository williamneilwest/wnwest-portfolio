import json
import os
from pathlib import Path
from datetime import datetime, timezone

from flask import Blueprint, current_app, jsonify, request

from ..models.platform import SessionLocal, SoftwareRegistry, init_platform_db
from ..services.authz import get_current_user, require_admin
from ..services.software_registry_service import save_software_registry, sync_software_registry_from_latest_file
from ..tools.document.csv_cleaner import normalize_headers, read_clean_table_rows
from ..utils.deprecation import log_deprecated_route

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - guarded at runtime for optional dependency
    load_workbook = None


software_bp = Blueprint('software', __name__)


REQUIRED_FIELDS = ('application_name',)
BOOLEAN_FIELDS = ('phi', 'corp_standard', 'baa')


def _utc_now():
    return datetime.now(timezone.utc)


def _normalize_cell(value) -> str:
    return ' '.join(str(value or '').strip().split())


def _normalize_bool_like(value) -> str:
    raw = _normalize_cell(value).lower()
    if raw in {'true', 'yes', 'y', '1'}:
        return 'Yes'
    if raw in {'false', 'no', 'n', '0'}:
        return 'No'
    if not raw:
        return ''
    return str(value or '').strip()


def _serialize_row(row: SoftwareRegistry) -> dict:
    return {
        'id': int(row.id),
        'vendor_name': row.vendor_name or '',
        'application_name': row.application_name or '',
        'business_function': row.business_function or '',
        'phi': row.phi or '',
        'corp_standard': row.corp_standard or '',
        'baa': row.baa or '',
        'core_level': row.core_level or '',
        'twilight': row.twilight or '',
        'system_owner': row.system_owner or '',
        'hosting_provider': row.hosting_provider or '',
        'deployed_sites': row.deployed_sites or '',
        'description': row.description or '',
        'business_owner': row.business_owner or '',
        'created_at': row.created_at.isoformat() if row.created_at else None,
        'updated_at': row.updated_at.isoformat() if row.updated_at else None,
    }


def _backend_data_dir() -> Path:
    configured = str(current_app.config.get('BACKEND_DATA_DIR') or '/app/data').strip() or '/app/data'
    return Path(configured)


def _software_storage_dir() -> Path:
    return _backend_data_dir() / 'software_registry'


def _latest_metadata_path() -> Path:
    return _software_storage_dir() / 'latest_approved_software.json'


def _utc_stamp() -> str:
    return _utc_now().strftime('%Y%m%dT%H%M%SZ')


def _read_latest_metadata() -> dict | None:
    metadata_path = _latest_metadata_path()
    if not metadata_path.exists():
        return None

    try:
        payload = json.loads(metadata_path.read_text(encoding='utf-8'))
    except Exception:
        return None

    return payload if isinstance(payload, dict) else None


def _write_latest_metadata(payload: dict) -> None:
    storage_dir = _software_storage_dir()
    storage_dir.mkdir(parents=True, exist_ok=True)
    _latest_metadata_path().write_text(json.dumps(payload, indent=2), encoding='utf-8')


def _extract_rows_from_xlsx(uploaded_stream) -> tuple[list[str], list[dict]]:
    if load_workbook is None:
        raise ValueError('XLSX support is unavailable until openpyxl is installed.')

    uploaded_stream.seek(0)
    workbook = load_workbook(filename=uploaded_stream, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows_iter = worksheet.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if header_row is None:
            return [], []

        headers = [str(cell or '').strip() for cell in header_row]
        payload_rows = []
        for row_values in rows_iter:
            payload_rows.append({
                headers[index]: '' if index >= len(row_values) or row_values[index] is None else str(row_values[index]).strip()
                for index in range(len(headers))
            })
        return headers, payload_rows
    finally:
        workbook.close()

def _save_latest_uploaded_file(uploaded_file, extension: str, total_rows: int, mode: str) -> dict:
    username = getattr(get_current_user(), 'username', None) or 'unknown'
    storage_dir = _software_storage_dir()
    archive_dir = storage_dir / 'archive'
    storage_dir.mkdir(parents=True, exist_ok=True)
    archive_dir.mkdir(parents=True, exist_ok=True)

    uploaded_file.stream.seek(0)
    latest_filename = f'latest_approved_software{extension}'
    latest_path = storage_dir / latest_filename
    uploaded_file.save(latest_path)

    uploaded_file.stream.seek(0)
    archive_filename = f'approved_software_{_utc_stamp()}{extension}'
    archive_path = archive_dir / archive_filename
    uploaded_file.save(archive_path)

    metadata = {
        'label': 'Latest Approved Software',
        'original_filename': str(uploaded_file.filename or '').strip() or latest_filename,
        'latest_filename': latest_filename,
        'latest_path': str(latest_path),
        'archive_filename': archive_filename,
        'archive_path': str(archive_path),
        'uploaded_by': username,
        'uploaded_at': _utc_now().isoformat(),
        'records': int(total_rows),
        'mode': mode,
        'source_type': extension.lstrip('.'),
    }
    _write_latest_metadata(metadata)
    return metadata


@software_bp.get('/api/software')
def list_software_registry():
    sync_result = sync_software_registry_from_latest_file(_software_storage_dir())
    init_platform_db()
    session = SessionLocal()
    try:
        rows = (
            session.query(SoftwareRegistry)
            .order_by(SoftwareRegistry.application_name.asc(), SoftwareRegistry.vendor_name.asc(), SoftwareRegistry.id.asc())
            .all()
        )
        return jsonify({
            'success': True,
            'items': [_serialize_row(row) for row in rows],
            'latest': _read_latest_metadata(),
            'source_sync': sync_result,
        })
    finally:
        session.close()


@software_bp.post('/api/software/upload')
@software_bp.post('/api/software-registry/upload')
def upload_software_registry():
    if request.path == '/api/software-registry/upload':
        log_deprecated_route('/api/software-registry/upload', '/api/software/upload')

    admin_error = require_admin()
    if admin_error is not None:
        return admin_error

    uploaded_file = request.files.get('file')
    if uploaded_file is None or not uploaded_file.filename:
        return jsonify({'success': False, 'error': 'A CSV, JSON, or XLSX file is required.'}), 400

    filename = str(uploaded_file.filename or '').strip()
    extension = os.path.splitext(filename.lower())[1]
    if extension not in {'.csv', '.json', '.xlsx'}:
        return jsonify({'success': False, 'error': 'File must be .csv, .json, or .xlsx.'}), 400

    if extension in {'.csv', '.json'}:
        try:
            headers, raw_rows = read_clean_table_rows(uploaded_file.stream, file_type=extension.lstrip('.'))
        except json.JSONDecodeError:
            return jsonify({'success': False, 'error': 'Invalid JSON payload.'}), 400
        except ValueError as exc:
            return jsonify({'success': False, 'error': str(exc)}), 400

        if not headers:
            empty_label = 'JSON' if extension == '.json' else 'CSV'
            return jsonify({'success': False, 'error': f'Empty {empty_label}'}), 400
    else:
        try:
            xlsx_headers, raw_xlsx_rows = _extract_rows_from_xlsx(uploaded_file.stream)
        except ValueError as exc:
            return jsonify({'success': False, 'error': str(exc)}), 400

        headers = normalize_headers(xlsx_headers)
        raw_rows = [[raw_row.get(source_header, '') for source_header in xlsx_headers] for raw_row in raw_xlsx_rows]

    if not headers:
        return jsonify({'success': False, 'error': 'File headers were not detected.'}), 400

    missing_headers = [field for field in REQUIRED_FIELDS if field not in set(headers)]
    if missing_headers:
        return jsonify(
            {
                'success': False,
                'error': f'Missing required headers for: {", ".join(missing_headers)}',
            }
        ), 400

    mode = str(request.args.get('mode') or request.form.get('mode') or 'replace').strip().lower()
    if mode not in {'replace', 'upsert'}:
        return jsonify({'success': False, 'error': 'mode must be "replace" or "upsert".'}), 400

    parsed_rows = []
    errors = []
    for row_index, row in enumerate(raw_rows, start=2):
        if len(row) != len(headers):
            errors.append(f'Row {row_index}: column mismatch')
            continue

        normalized = {}
        for index, header in enumerate(headers):
            cell_value = row[index]
            normalized[header] = _normalize_bool_like(cell_value) if header in BOOLEAN_FIELDS else _normalize_cell(cell_value)

        missing_required = [field for field in REQUIRED_FIELDS if not _normalize_cell(normalized.get(field, ''))]
        if missing_required:
            errors.append(f'Row {row_index}: missing {", ".join(missing_required)}')
            continue

        parsed_rows.append(normalized)

    if not parsed_rows:
        return jsonify({'success': False, 'error': 'No valid records to save.', 'errors': errors[:200], 'preview': []}), 400

    save_result = save_software_registry(parsed_rows, mode=mode)
    latest = _save_latest_uploaded_file(uploaded_file, extension, int(save_result.get('total', 0)), mode)
    return jsonify({
        'success': True,
        'status': 'success',
        'mode': mode,
        'inserted': int(save_result.get('inserted', 0)),
        'updated': int(save_result.get('updated', 0)),
        'total': int(save_result.get('total', 0)),
        'records_saved': len(parsed_rows),
        'errors': errors[:200],
        'preview': parsed_rows[:10],
        'latest': latest,
    })


@software_bp.get('/api/software/search')
def search_software_registry():
    init_platform_db()
    session = SessionLocal()
    try:
        query = str(request.args.get('q') or '').strip()
        limit_raw = str(request.args.get('limit') or '50').strip()
        try:
            limit = max(1, min(int(limit_raw), 200))
        except ValueError:
            limit = 50

        rows_query = session.query(SoftwareRegistry)
        if query:
            rows_query = rows_query.filter(SoftwareRegistry.application_name.ilike(f'%{query}%'))

        selected = (
            rows_query
            .order_by(SoftwareRegistry.application_name.asc(), SoftwareRegistry.vendor_name.asc(), SoftwareRegistry.id.asc())
            .limit(limit)
            .all()
        )

        return jsonify({
            'success': True,
            'query': query.lower(),
            'limit': limit,
            'count': len(selected),
            'items': [_serialize_row(row) for row in selected],
            'latest': _read_latest_metadata(),
        })
    finally:
        session.close()
