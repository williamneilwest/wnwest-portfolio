import csv
import hashlib
import io
import re
from pathlib import Path

from ...utils.normalizers.base_normalizer import normalize_row
from ...utils.normalizers.hardware_normalizer import normalize_hardware_record

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None


def normalize_source_name(value) -> str:
    raw = str(value or '').strip().lower()
    return '_'.join(raw.split())


def normalize_source_header(value) -> str:
    text = str(value or '').strip().lower()
    if not text:
        return ''

    normalized = re.sub(r'[^a-z0-9]+', '_', text)
    normalized = re.sub(r'_+', '_', normalized).strip('_')
    return normalized


def normalize_source_row_headers(row: dict) -> dict:
    normalized = {}
    if not isinstance(row, dict):
        return normalized

    for key, value in row.items():
        normalized_key = normalize_source_header(key)
        if not normalized_key:
            continue
        normalized[normalized_key] = value

    return normalized


def read_rows_from_csv(path: str) -> list[dict]:
    source = Path(str(path or '')).expanduser()
    if not source.exists() or not source.is_file():
        return []

    decoded = source.read_bytes().decode('utf-8-sig', errors='replace')
    reader = csv.DictReader(io.StringIO(decoded))
    return [normalize_source_row_headers(dict(row or {})) for row in reader]


def read_rows_from_excel(path: str) -> list[dict]:
    source = Path(str(path or '')).expanduser()
    if not source.exists() or not source.is_file() or load_workbook is None:
        return []

    workbook = load_workbook(filename=str(source), read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows_iter = worksheet.iter_rows(values_only=True)
        try:
          headers = next(rows_iter)
        except StopIteration:
          return []

        normalized_headers = [normalize_source_header(header) for header in headers]
        if not any(normalized_headers):
            return []

        records = []
        for values in rows_iter:
            row = {}
            has_value = False
            for index, header in enumerate(normalized_headers):
                key = header or f'column_{index + 1}'
                cell_value = values[index] if values and index < len(values) else None
                if cell_value is not None and str(cell_value).strip() != '':
                    has_value = True
                row[key] = '' if cell_value is None else str(cell_value)
            if has_value:
                records.append(normalize_source_row_headers(row))
        return records
    finally:
        workbook.close()


def read_rows(path: str) -> list[dict]:
    source = Path(str(path or '')).expanduser()
    extension = source.suffix.lower()
    if extension in {'.xlsx', '.xlsm', '.xltx', '.xltm'}:
        return read_rows_from_excel(path)
    return read_rows_from_csv(path)


def compute_file_checksum(path: str) -> str:
    source = Path(str(path or '')).expanduser()
    if not source.exists() or not source.is_file():
        return ''

    digest = hashlib.sha256()
    with source.open('rb') as handle:
        for chunk in iter(lambda: handle.read(8192), b''):
            digest.update(chunk)
    return digest.hexdigest()


def count_rows(path: str) -> int:
    return len(read_rows(path))


def normalize(source_name, data):
    source_key = normalize_source_name(source_name)
    rows = data if isinstance(data, list) else []

    if source_key in {'hardware_inventory', 'ref_hardware', 'hardware_rmr'}:
        if rows:
            print("RAW:", rows[0])
            print("NORMALIZED:", normalize_hardware_record(rows[0]))
        return [normalize_hardware_record(row) for row in rows]

    # Generic key normalization for all other sources.
    return [normalize_row(row) for row in rows]
